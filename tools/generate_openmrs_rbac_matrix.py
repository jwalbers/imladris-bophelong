"""
generate_openmrs_rbac_matrix.py — Build the OpenMRS Mirebalais/PIH EMR RBAC matrix.

Reads role and privilege CSVs from the openmrs-config-zl repo (expected as a
sibling of imladris-bophelong) and writes openmrs-rbac-matrix.xlsx.

Source files (relative to imladris/ repo root):
  openmrs/openmrs-config-zl/configuration/roles/roles_haiti.csv
  openmrs/openmrs-config-zl/configuration/roles/roles_haiti_hiv.csv
  openmrs/openmrs-config-zl/configuration/roles/roles_legacy.csv
  openmrs/openmrs-config-zl/content/target/parent/configuration/
      backend_configuration/roles/roles.csv
  openmrs/openmrs-config-zl/content/target/parent/configuration/
      backend_configuration/privileges/privileges.csv

Output: openmrs-rbac-matrix.xlsx  (written to current directory)

Run from imladris-bophelong/:
  python tools/generate_openmrs_rbac_matrix.py [--imladris-root PATH]
"""

import argparse
import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def var_to_priv(var: str) -> str | None:
    """Convert ${privilege.app_foo_bar} → 'App: foo.barBaz' (best-effort)."""
    m = re.match(r'\$\{privilege\.(app|task)_(.+)\}', var)
    if not m:
        return None
    kind = "App" if m.group(1) == "app" else "Task"
    parts = m.group(2).split("_")
    module = parts[0]
    rest = parts[1:]
    suffix = rest[0] + "".join(p.capitalize() for p in rest[1:]) if rest else ""
    return f"{kind}: {module}.{suffix}" if suffix else f"{kind}: {module}"


def load_privileges(priv_csv: Path) -> dict[str, str]:
    """Return mapping of actual_privilege_string → description."""
    mapping = {}
    with open(priv_csv) as f:
        for row in csv.DictReader(f):
            actual = var_to_priv(row["Privilege name"].strip())
            if actual:
                mapping[actual] = row["Description"].strip()
    return mapping


def load_roles(role_files: list[Path]) -> dict[str, set[str]]:
    """Return mapping of role_name → set of privilege strings."""
    roles: dict[str, set[str]] = {}
    for fpath in role_files:
        with open(fpath) as f:
            for row in csv.DictReader(f):
                name = row["Role name"].strip().replace("Application Role: ", "")
                privs = {p.strip() for p in row.get("Privileges", "").split(";") if p.strip()}
                roles.setdefault(name, set()).update(privs)
    return roles


def build_xlsx(roles: dict, priv_desc: dict, out_path: Path) -> None:
    all_privs = sorted({p for ps in roles.values() for p in ps})
    ordered = (
        sorted(p for p in all_privs if p.startswith("App:")) +
        sorted(p for p in all_privs if p.startswith("Task:")) +
        sorted(p for p in all_privs if not p.startswith(("App:", "Task:")))
    )
    role_names = sorted(roles.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RBAC Matrix"

    HDR   = PatternFill("solid", fgColor="1F3864")
    APP   = PatternFill("solid", fgColor="2E4057")
    TASK  = PatternFill("solid", fgColor="1B3A2D")
    OTHER = PatternFill("solid", fgColor="3A1C1C")
    XFILL = PatternFill("solid", fgColor="D9EAD3")
    ALT   = PatternFill("solid", fgColor="F8F8F8")
    WHITE = PatternFill("solid", fgColor="FFFFFF")
    thin  = Side(style="thin", color="CCCCCC")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col, val in [(1, "Privilege"), (2, "Description")]:
        c = ws.cell(1, col, val)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for ci, rname in enumerate(role_names, 3):
        c = ws.cell(1, ci, rname)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = HDR
        c.alignment = Alignment(text_rotation=60, vertical="bottom", horizontal="center")

    section_map = {"App:": ("── App Privileges ──", APP),
                   "Task:": ("── Task Privileges ──", TASK),
                   "Other": ("── Other Privileges ──", OTHER)}
    current = None
    ri = 2

    for priv in ordered:
        section = "App:" if priv.startswith("App:") else ("Task:" if priv.startswith("Task:") else "Other")
        if section != current:
            current = section
            label, fill = section_map[section]
            c = ws.cell(ri, 1, label)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = fill
            ws.merge_cells(start_row=ri, start_column=1,
                           end_row=ri, end_column=2 + len(role_names))
            ri += 1

        row_fill = ALT if ri % 2 == 0 else WHITE
        short = priv.replace("App: ", "").replace("Task: ", "")
        c = ws.cell(ri, 1, short)
        c.font = Font(size=9); c.fill = row_fill
        c.alignment = Alignment(vertical="center"); c.border = bdr

        c = ws.cell(ri, 2, priv_desc.get(priv, ""))
        c.font = Font(size=8, italic=True, color="555555"); c.fill = row_fill
        c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = bdr

        for ci, rname in enumerate(role_names, 3):
            if priv in roles[rname]:
                c = ws.cell(ri, ci, "x")
                c.font = Font(bold=True, size=9, color="1A6B2E"); c.fill = XFILL
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c = ws.cell(ri, ci, ""); c.fill = row_fill
            c.border = bdr
        ri += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 48
    for ci in range(3, 3 + len(role_names)):
        ws.column_dimensions[get_column_letter(ci)].width = 6
    ws.row_dimensions[1].height = 110
    ws.freeze_panes = "C2"
    wb.save(out_path)
    print(f"Saved: {out_path}  ({out_path.stat().st_size // 1024} KB)  "
          f"[{len(role_names)} roles × {len(ordered)} privileges]")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--imladris-root", default="../imladris",
                        help="Path to the imladris repo root")
    args = parser.parse_args()

    root = Path(args.imladris_root) / "openmrs/openmrs-config-zl"
    role_files = [
        root / "configuration/roles/roles_haiti.csv",
        root / "configuration/roles/roles_haiti_hiv.csv",
        root / "configuration/roles/roles_legacy.csv",
        root / "content/target/parent/configuration/backend_configuration/roles/roles.csv",
    ]
    priv_csv = root / "content/target/parent/configuration/backend_configuration/privileges/privileges.csv"

    priv_desc = load_privileges(priv_csv)
    roles = load_roles(role_files)
    build_xlsx(roles, priv_desc, Path("openmrs-rbac-matrix.xlsx"))


if __name__ == "__main__":
    main()
